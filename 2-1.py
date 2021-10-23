import tkinter.messagebox
from tkinter import Tk, Menu, Toplevel, Label, Entry, StringVar, Radiobutton, IntVar, Button, Listbox, END
from tkinter.ttk import Combobox, Checkbutton
from openpyxl import *

my_window = Tk()

my_window.title('Menu Test')

my_window.geometry('800x250+500+300')
menu_bar = Menu(my_window)

stu_menu = Menu(menu_bar)
score_menu = Menu(menu_bar)

Label(my_window, text='欢迎来到、\n智能学生个人信息考试等级格式化系统\n使用前,请关闭编辑软件', width=100, font='Arial,200').place(x=10, y=50)


def write_sport_data_head():
    work_book = load_workbook('stu.xlsx')
    sheet = work_book.active
    sheet.title = 'student info'

    sheet['A1'] = '姓名'
    sheet['B1'] = '班级'
    sheet['C1'] = '性别'
    sheet['D1'] = '年龄'
    sheet['E1'] = '身高'
    sheet['F1'] = '体重'
    sheet['G1'] = '籍贯'
    sheet['H1'] = '足球'
    sheet['I1'] = '篮球'
    sheet['J1'] = '乒乓球'
    sheet['K1'] = '羽毛球'
    sheet['L1'] = '跑步'
    sheet['M1'] = '跳绳'
    sheet['N1'] = '游泳'
    sheet['O1'] = '其他'
    sheet['P1'] = '平均运动时间'
    work_book.save(filename='stu.xlsx')


def write_sport_data_content(data_list: list):
    work_book = load_workbook('stu.xlsx')
    sheet = work_book.active
    max_row = sheet.max_row
    row_to_write = max_row + 1
    for data_column in range(1, 17):
        sheet.cell(row=row_to_write, column=data_column, value=data_list[data_column - 1])
    work_book.save(filename='stu.xlsx')
    work_book.close()


def pop_add_stu_sport_win():
    stu_add_win = Toplevel()
    stu_add_win.title('添加学生信息')
    stu_add_win.geometry('600x480')

    Label(stu_add_win, text='姓名', width=10).place(x=10, y=20)
    stu_name = StringVar()
    Entry(stu_add_win, textvariable=stu_name, width=15).place(x=100, y=20)

    Label(stu_add_win, text='学号', width=10).place(x=10, y=50)
    stu_number = StringVar()
    Entry(stu_add_win, textvariable=stu_number, width=15).place(x=100, y=50)

    Label(stu_add_win, text='性别', width=10).place(x=10, y=80)
    gender_text = StringVar()
    Radiobutton(stu_add_win, text='男', variable=gender_text, value='男').place(x=100, y=80)
    Radiobutton(stu_add_win, text='女', variable=gender_text, value='女').place(x=180, y=80)

    Label(stu_add_win, text='年龄', width=10).place(x=10, y=110)
    age_number = StringVar()
    Entry(stu_add_win, textvariable=age_number, width=15).place(x=100, y=110)
    Label(stu_add_win, text='岁', width=2).place(x=210, y=110)

    Label(stu_add_win, text='身高', width=10).place(x=10, y=140)
    height_number = StringVar()
    Entry(stu_add_win, textvariable=height_number, width=15).place(x=100, y=140)
    Label(stu_add_win, text='厘米', width=4).place(x=210, y=140)

    Label(stu_add_win, text='体重', width=10).place(x=10, y=170)
    weight_number = StringVar()
    Entry(stu_add_win, textvariable=weight_number, width=15).place(x=100, y=170)
    Label(stu_add_win, text='公斤', width=4).place(x=210, y=170)

    Label(stu_add_win, text='籍贯', width=10).place(x=10, y=200)
    hometown = StringVar()
    province_value = ['广东', '广西', '湖南', '湖北', '江西', '福建', '海南', '云南', '贵州', '浙江', '江苏', '上海', '北京', '重庆']
    Combobox(stu_add_win, textvariable=hometown, width=12, values=province_value, state='readonly').place(x=100, y=200)

    Label(stu_add_win, text='运动爱好', width=10).place(x=10, y=230)
    hobby_sport1 = IntVar()
    hobby_sport2 = IntVar()
    hobby_sport3 = IntVar()
    hobby_sport4 = IntVar()
    hobby_sport5 = IntVar()
    hobby_sport6 = IntVar()
    hobby_sport7 = IntVar()
    hobby_sport8 = IntVar()

    Checkbutton(stu_add_win, text='足球', variable=hobby_sport1, onvalue=1, offvalue=0).place(x=50, y=260)
    Checkbutton(stu_add_win, text='篮球', variable=hobby_sport2, onvalue=1, offvalue=0).place(x=130, y=260)
    Checkbutton(stu_add_win, text='乒乓球', variable=hobby_sport3, onvalue=1, offvalue=0).place(x=210, y=260)
    Checkbutton(stu_add_win, text='羽毛球', variable=hobby_sport4, onvalue=1, offvalue=0).place(x=290, y=260)
    Checkbutton(stu_add_win, text='跑步', variable=hobby_sport5, onvalue=1, offvalue=0).place(x=50, y=290)
    Checkbutton(stu_add_win, text='跳绳', variable=hobby_sport6, onvalue=1, offvalue=0).place(x=130, y=290)
    Checkbutton(stu_add_win, text='游泳', variable=hobby_sport7, onvalue=1, offvalue=0).place(x=210, y=290)
    Checkbutton(stu_add_win, text='其他', variable=hobby_sport8, onvalue=1, offvalue=0).place(x=290, y=290)

    Label(stu_add_win, text='平均运动时长', width=10).place(x=20, y=320)
    sport_hours_number = StringVar()
    Entry(stu_add_win, textvariable=sport_hours_number, width=15).place(x=100, y=320)
    Label(stu_add_win, text='小时每周', width=10).place(x=200, y=320)

    def write_sport_excel_file():
        name_value = stu_name.get()
        stu_number_value = stu_number.get()
        gender_value = gender_text.get()
        age_value = age_number.get()
        height_value = height_number.get()
        weight_value = weight_number.get()
        hometown_value = hometown.get()
        hobby_sport1_value = str(hobby_sport1.get())
        hobby_sport2_value = str(hobby_sport2.get())
        hobby_sport3_value = str(hobby_sport3.get())
        hobby_sport4_value = str(hobby_sport4.get())
        hobby_sport5_value = str(hobby_sport5.get())
        hobby_sport6_value = str(hobby_sport6.get())
        hobby_sport7_value = str(hobby_sport7.get())
        hobby_sport8_value = str(hobby_sport8.get())
        sport_hours_value = sport_hours_number.get()

        stu_data = [name_value, stu_number_value, gender_value, age_value, height_value, weight_value,
                    hometown_value, hobby_sport1_value, hobby_sport2_value, hobby_sport3_value, hobby_sport4_value,
                    hobby_sport5_value, hobby_sport6_value, hobby_sport7_value, hobby_sport8_value, sport_hours_value]

        write_sport_data_content(stu_data)
        tkinter.messagebox.showinfo(title='成功', message='添加成功')

    Button(stu_add_win, text='确定添加', command=write_sport_excel_file, width=20).place(x=200, y=400)


def write_grades_head():
    work_book = Workbook()
    sheet = work_book.active
    sheet.title = '成绩等级表'

    sheet['A1'] = '姓名'
    sheet['B1'] = '学号'
    sheet['C1'] = '语文考试等级'
    sheet['D1'] = '数学考试等级'
    sheet['E1'] = '英语考试等级'
    sheet['F1'] = '科学考试等级'
    sheet['G1'] = '综合考试等级'
    sheet['H1'] = '语文考试分数'
    sheet['I1'] = '数学考试分数'
    sheet['J1'] = '英语考试分数'
    sheet['K1'] = '科学考试分数'
    sheet['L1'] = '平均考试分数'
    sheet['M1'] = '综合考试分数'

    work_book.save(filename='grades.xlsx')
    work_book.close()


def write_grades_content(data_list: list):
    work_book = load_workbook('grades.xlsx')
    sheet = work_book.active
    max_row = sheet.max_row
    row_to_write = max_row + 1
    for data_column in range(1, 14):
        sheet.cell(row=row_to_write, column=data_column, value=data_list[data_column - 1])
    work_book.save(filename='grades.xlsx')
    work_book.close()


def pop_add_stu_grades_win():
    stu_add_grades_win = Toplevel()
    stu_add_grades_win.title('添加学生等级')
    stu_add_grades_win.geometry('600x300')

    Label(stu_add_grades_win, text='姓名', width=10).place(x=10, y=20)
    name = StringVar()
    Entry(stu_add_grades_win, textvariable=name, width=15).place(x=100, y=20)

    Label(stu_add_grades_win, text='班级', width=10).place(x=10, y=50)
    class_name = StringVar()
    Entry(stu_add_grades_win, textvariable=class_name, width=15).place(x=100, y=50)

    Label(stu_add_grades_win, text='语文考试等级', width=10).place(x=10, y=80)
    grades_chi = StringVar()
    Entry(stu_add_grades_win, textvariable=grades_chi, width=15).place(x=100, y=80)

    Label(stu_add_grades_win, text='数学考试等级', width=10).place(x=10, y=110)
    grades_mat = StringVar()
    Entry(stu_add_grades_win, textvariable=grades_mat, width=15).place(x=100, y=110)

    Label(stu_add_grades_win, text='英语考试等级', width=10).place(x=10, y=140)
    grades_eng = StringVar()
    Entry(stu_add_grades_win, textvariable=grades_eng, width=15).place(x=100, y=140)

    Label(stu_add_grades_win, text='科学考试等级', width=10).place(x=10, y=170)
    grades_sci = StringVar()
    Entry(stu_add_grades_win, textvariable=grades_sci, width=15).place(x=100, y=170)

    Label(stu_add_grades_win, text='综合等级', width=10).place(x=10, y=220)
    total_grades = StringVar()
    Entry(stu_add_grades_win, textvariable=total_grades, width=15).place(x=100, y=220)

    Label(stu_add_grades_win, text='学号 ', width=10).place(x=290, y=20)
    stu_number = StringVar()
    Entry(stu_add_grades_win, textvariable=stu_number, width=15).place(x=380, y=20)

    Label(stu_add_grades_win, text='单元', width=10).place(x=290, y=50)
    unit_score = StringVar()
    Entry(stu_add_grades_win, textvariable=unit_score, width=15).place(x=380, y=50)

    Label(stu_add_grades_win, text='语文考试分数', width=10).place(x=290, y=80)
    chi_score = StringVar()
    Entry(stu_add_grades_win, textvariable=chi_score, width=15).place(x=380, y=80)

    Label(stu_add_grades_win, text='数学考试分数', width=10).place(x=290, y=110)
    mat_score = StringVar()
    Entry(stu_add_grades_win, textvariable=mat_score, width=15).place(x=380, y=110)

    Label(stu_add_grades_win, text='英语考试分数', width=10).place(x=290, y=140)
    eng_score = StringVar()
    Entry(stu_add_grades_win, textvariable=eng_score, width=15).place(x=380, y=140)

    Label(stu_add_grades_win, text='科学考试分数', width=10).place(x=290, y=170)
    sci_score = StringVar()
    Entry(stu_add_grades_win, textvariable=sci_score, width=15).place(x=380, y=170)

    Label(stu_add_grades_win, text='综合分数', width=10).place(x=290, y=220)
    total_score = StringVar()
    Entry(stu_add_grades_win, textvariable=total_score, width=15).place(x=380, y=220)

    def write_grades_excel_file():
        name_value = name.get()
        class_name_value = class_name.get()
        stu_number_value = stu_number.get()
        unit_score_value = unit_score.get()
        chi_grades_value = grades_chi.get()
        mat_grades_value = grades_mat.get()
        eng_grades_value = grades_eng.get()
        sci_grades_value = grades_sci.get()
        total_grades_value = total_grades.get()
        chi_score_value = chi_score.get()
        mat_score_value = mat_score.get()
        eng_score_value = eng_score.get()
        sci_score_value = sci_score.get()
        total_score_value = total_score.get()

        grades_data = [name_value, class_name_value, stu_number_value, unit_score_value, chi_grades_value,
                       mat_grades_value, eng_grades_value, sci_grades_value, total_grades_value, chi_score_value,
                       mat_score_value, eng_score_value, sci_score_value, total_score_value]

        write_grades_content(grades_data)
        tkinter.messagebox.showinfo(title='成功', message='添加成功')

    Button(stu_add_grades_win, text='确认添加', command=write_grades_excel_file, width=20).place(x=150, y=250)


def pop_query_stu_grades_win():
    stu_query_score_win = Toplevel()
    stu_query_score_win.title('学生个人信息')
    stu_query_score_win.geometry('800x480')

    def read_excel_file() -> list:
        work_book = load_workbook('grades.xlsx')
        sheet = work_book.active
        stu_score_info_list = []
        for row in sheet.iter_rows(min_col=1, max_col=14, min_row=1, max_row=sheet.max_row):
            score_info_temp_list = []
            for c in row:
                score_info_temp_list.append(c.value)

            stu_score_info_list.append(score_info_temp_list)
        work_book.close()

        return stu_score_info_list

    score_query_return_list = read_excel_file()

    list_value = StringVar()
    listbox_sport_info = Listbox(stu_query_score_win, listvariable=list_value, width=100)
    for i in score_query_return_list:
        listbox_sport_info.insert(END, i)

    listbox_sport_info.pack(side='top')


def pop_query_stu_sport_win():
    stu_query_sport_win = Toplevel()
    stu_query_sport_win.title('学生运动与体格信息')
    stu_query_sport_win.geometry('800x480')

    def read_excel_file() -> list:
        work_book = load_workbook('stu.xlsx')
        sheet = work_book.active
        stu_sport_info_list = []
        for row in sheet.iter_rows(min_col=1, max_col=16, min_row=1, max_row=sheet.max_row):
            sport_info_temp_list = []
            for c in row:
                sport_info_temp_list.append(c.value)

            stu_sport_info_list.append(sport_info_temp_list)
        work_book.close()

        return stu_sport_info_list

    sport_query_return_list = read_excel_file()

    list_value = StringVar()
    listbox_sport_info = Listbox(stu_query_sport_win, listvariable=list_value, width=100)
    for i in sport_query_return_list:
        listbox_sport_info.insert(END, i)

    listbox_sport_info.pack(side='top')


stu_menu.add_command(label='学生信息', command='')
stu_menu.add_command(label='阅览登记', command='')
stu_menu.add_command(label='删除', command='')
menu_bar.add_cascade(label='记录', menu=stu_menu)

stu_menu = Menu(menu_bar)

score_menu.add_command(label='表格', command='')
score_menu.add_command(label='文本', command='')
stu_menu.add_command(label='高级操作', command='')
menu_bar.add_cascade(label='生成', menu=score_menu)

score_menu = Menu(menu_bar)

my_window.config(menu=menu_bar)

my_window.mainloop()
